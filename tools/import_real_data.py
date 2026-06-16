import argparse
import json
import re
import sys
import urllib.request
import urllib.parse
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl


API_URL = "https://script.google.com/macros/s/AKfycbx_xHAyop1J6QbSez84NWqZ5Ld1xEmCyZW0HJqFLijVGzjk6URMtbs_OPujQ4QXXPk/exec"


def cell_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def cell_month(value):
    parsed = cell_date(value)
    return parsed[:7] + "-01" if parsed else ""


def cell_time(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if not text:
        return ""
    if " " in text:
        try:
            return datetime.strptime(text, "%Y-%m-%d %H:%M:%S").time().strftime("%H:%M")
        except ValueError:
            pass
    match = re.match(r"^(\d{1,2}):(\d{2})", text)
    return f"{int(match.group(1)):02d}:{match.group(2)}" if match else text


def number(value, default=""):
    if value in (None, ""):
        return default
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return default


def int_id(value):
    if value in (None, ""):
        return ""
    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return str(value).strip()


def clean_text(value):
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def contact(value):
    value = clean_text(value)
    if not value:
        return ""
    try:
        return str(int(float(value)))
    except ValueError:
        return value


def client_code(name):
    clean = re.sub(r"[^A-Za-z0-9 ]+", " ", name).strip()
    words = [w for w in clean.split() if w.lower() not in {"pvt", "ltd", "private", "limited"}]
    if not words:
        return "CL"
    if len(words) == 1:
        return words[0][:3].upper()
    return "".join(word[0] for word in words[:3]).upper()


def make_id(*parts):
    text = "-".join(clean_text(part) for part in parts if clean_text(part))
    return re.sub(r"[^A-Za-z0-9-]+", "-", text).strip("-").upper()


def rows_after_header(ws, header_name):
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if clean_text(row[0]) == header_name:
            headers = [clean_text(v) for v in row]
            for data_row in ws.iter_rows(min_row=idx + 1, values_only=True):
                if not any(v not in (None, "") for v in data_row):
                    continue
                yield headers, data_row


def parse_employees(wb):
    ws = wb["Employees"]
    records = []
    for headers, row in rows_after_header(ws, "Employee ID"):
        emp_id = int_id(row[0])
        name = clean_text(row[1])
        if not emp_id or not name:
            continue
        records.append({
            "EmployeeID": emp_id,
            "EmployeeName": name,
            "Contact": contact(row[2]),
            "Department": clean_text(row[3]),
            "Designation": clean_text(row[4]),
            "SalaryType": "Monthly",
            "MonthlySalary": number(row[5]),
            "PerDaySalary": number(row[6]),
            "OTRate": number(row[7]),
            "Status": "Active",
        })
    return records


def parse_advances(wb):
    ws = wb["Advance Register"]
    records = []
    counter = 1
    for headers, row in rows_after_header(ws, "Date"):
        emp_id = int_id(row[1])
        amount = number(row[3])
        row_date = cell_date(row[0])
        if not row_date or not emp_id or amount == "":
            continue
        remarks = clean_text(row[6])
        mode = clean_text(row[4])
        if mode:
            remarks = f"Payment Mode: {mode}" + (f"; {remarks}" if remarks and remarks != "-" else "")
        records.append({
            "AdvanceID": make_id("ADV", row_date.replace("-", ""), emp_id, counter),
            "EmployeeID": emp_id,
            "Date": row_date,
            "Amount": amount,
            "Status": clean_text(row[5]) or "Pending",
            "Remarks": "" if remarks == "-" else remarks,
        })
        counter += 1
    return records


def parse_salary_register(wb):
    ws = wb["Salary Register"]
    records = []
    counter = 1
    for headers, row in rows_after_header(ws, "Month"):
        month = cell_month(row[0])
        emp_id = int_id(row[1])
        if not month or not emp_id:
            continue
        records.append({
            "SalaryID": make_id("SAL", month[:7], emp_id, counter),
            "Month": month,
            "EmployeeID": emp_id,
            "EmployeeName": clean_text(row[2]),
            "GrossSalary": number(row[3]),
            "AdvanceDeducted": number(row[4], 0),
            "OtherDeductions": number(row[5], 0),
            "NetSalaryPayable": number(row[6]),
            "PaymentStatus": clean_text(row[7]) or "Pending",
            "PaymentDate": cell_date(row[8]),
        })
        counter += 1
    return records


def parse_attendance(wb):
    records = []
    counter = 1
    for ws in wb.worksheets:
        if not re.match(r"^\d{2}-2026$", ws.title):
            continue
        in_attendance = False
        current_date = ""
        for row in ws.iter_rows(values_only=True):
            if clean_text(row[0]) == "Date" and clean_text(row[1]) == "Employee ID":
                in_attendance = True
                continue
            if not in_attendance:
                continue
            row_date = cell_date(row[0]) or current_date
            if cell_date(row[0]):
                current_date = row_date
            emp_id = int_id(row[1])
            emp_name = clean_text(row[2])
            if not row_date or not emp_id or not emp_name or emp_name == "#N/A":
                continue
            work_hours = number(row[11], 0)
            ot_hours = parse_hours(row[10])
            records.append({
                "AttendanceID": make_id("ATT", row_date.replace("-", ""), emp_id, counter),
                "EmployeeID": emp_id,
                "Date": row_date,
                "CheckIn": cell_time(row[3]),
                "BreakStart": cell_time(row[4]),
                "BreakEnd": cell_time(row[5]),
                "CheckOut": cell_time(row[6]),
                "WorkMinutes": int(round(float(work_hours or 0) * 60)),
                "OTHours": ot_hours,
                "Remarks": f"Source: {ws.title}; Employee: {emp_name}",
            })
            counter += 1
    return records


def parse_hours(value):
    if value in (None, ""):
        return 0
    if isinstance(value, datetime):
        return round(value.hour + value.minute / 60, 2)
    if isinstance(value, time):
        return round(value.hour + value.minute / 60, 2)
    text = str(value).strip()
    if re.match(r"^\d{1,2}:\d{2}", text):
        hh, mm = text.split(":")[:2]
        return round(int(hh) + int(mm) / 60, 2)
    try:
        return round(float(text), 2)
    except ValueError:
        return 0


def parse_invoices(wb):
    ws = wb.active
    clients = {}
    invoices = []
    payments = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        invoice_no = clean_text(row[0])
        company = clean_text(row[2])
        amount = number(row[3])
        if not invoice_no or not company or amount == "":
            continue
        code = client_code(company)
        clients[code] = {
            "ClientCode": code,
            "ClientName": company,
            "GSTIN": "",
            "Address": "",
            "StateCode": "24",
            "Status": "Active",
        }
        payment_status = clean_text(row[4]) or "Pending"
        payment_date = cell_date(row[5])
        invoice_date = cell_date(row[1])
        invoices.append({
            "InvoiceID": make_id("INV", invoice_no),
            "InvoiceNo": invoice_no,
            "OrderID": "",
            "ClientCode": code,
            "InvoiceDate": invoice_date,
            "Amount": amount,
            "DueDate": due_date(invoice_date, 30),
            "PaymentStatus": "Paid" if payment_status.lower().startswith("paid") else "Unpaid",
            "PaymentDate": payment_date,
        })
        if payment_status.lower().startswith("paid"):
            payments.append({
                "PaymentID": make_id("PAY", invoice_no),
                "InvoiceID": make_id("INV", invoice_no),
                "ClientCode": code,
                "AmountReceived": amount,
                "PaymentDate": payment_date,
                "PaymentMode": clean_text(row[6]) or "Bank",
                "ReferenceNo": "",
                "Remarks": "Imported from invoice payment sheet",
            })
    return list(clients.values()), invoices, payments


def due_date(value, days):
    parsed = cell_date(value)
    if not parsed:
        return ""
    return (datetime.strptime(parsed, "%Y-%m-%d").date() + timedelta(days=days)).isoformat()


def build_payload(source_1, source_2):
    payroll_wb = openpyxl.load_workbook(source_1, data_only=True, read_only=False)
    invoice_wb = openpyxl.load_workbook(source_2, data_only=True, read_only=False)
    clients, invoices, customer_payments = parse_invoices(invoice_wb)
    return {
        "Clients": clients,
        "Employees": parse_employees(payroll_wb),
        "Advances": parse_advances(payroll_wb),
        "SalaryRegister": parse_salary_register(payroll_wb),
        "Attendance": parse_attendance(payroll_wb),
        "Invoices": invoices,
        "CustomerPayments": customer_payments,
    }


def post_json(api_url, payload):
    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(api_url, data=data, headers={"Content-Type": "text/plain;charset=utf-8"}, method="POST")
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def get_json(api_url, params):
    url = api_url + "?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def request_with_retry(fn, attempts=4):
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return fn()
        except Exception as error:
            last_error = error
            if attempt == attempts:
                break
    raise last_error


def upload_payload(api_url, username, password, payload):
    login = request_with_retry(lambda: post_json(api_url, {"action": "login", "data": {"username": username, "password": password}}))
    token = login.get("data", {}).get("token")
    if not token:
        raise RuntimeError(f"Login failed: {login}")
    results = []
    action_names = {
        "Clients": "saveClients",
        "Employees": "saveEmployees",
        "Advances": "saveAdvances",
        "SalaryRegister": "saveSalaryRegister",
        "Attendance": "saveAttendance",
        "Invoices": "saveInvoices",
        "CustomerPayments": "saveCustomerPayments",
    }
    get_actions = {
        "Clients": ("getClients", "ClientCode"),
        "Employees": ("getEmployees", "EmployeeID"),
        "Advances": ("getAdvances", "AdvanceID"),
        "SalaryRegister": ("getSalaryRegister", "SalaryID"),
        "Attendance": ("getAttendance", "AttendanceID"),
        "Invoices": ("getInvoices", "InvoiceID"),
        "CustomerPayments": ("getCustomerPayments", "PaymentID"),
    }
    for sheet, rows in payload.items():
        action = action_names[sheet]
        get_action, id_field = get_actions[sheet]
        existing_response = request_with_retry(lambda: get_json(api_url, {"action": get_action, "token": token}))
        existing_ids = {str(row.get(id_field, "")) for row in existing_response.get("data", [])}
        for row in rows:
            row_id = str(row.get(id_field, ""))
            if row_id in existing_ids:
                results.append({"sheet": sheet, "id": row_id, "success": True, "message": "Skipped existing"})
                continue
            result = request_with_retry(lambda: post_json(api_url, {"action": action, "token": token, "data": row}))
            results.append({"sheet": sheet, "id": row_id, "success": result.get("success"), "message": result.get("message")})
            if not result.get("success"):
                raise RuntimeError(json.dumps(results[-1], indent=2))
    return results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-1", default="imports/source-1.xlsx")
    parser.add_argument("--source-2", default="imports/source-2.xlsx")
    parser.add_argument("--output", default="imports/erp-import-payload.json")
    parser.add_argument("--upload", action="store_true")
    parser.add_argument("--api-url", default=API_URL)
    parser.add_argument("--username", default="admin")
    parser.add_argument("--password", default="admin123")
    args = parser.parse_args()

    payload = build_payload(args.source_1, args.source_2)
    Path(args.output).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    summary = {sheet: len(rows) for sheet, rows in payload.items()}
    print(json.dumps({"output": args.output, "summary": summary}, indent=2))
    if args.upload:
        results = upload_payload(args.api_url, args.username, args.password, payload)
        print(json.dumps({"uploaded": len(results)}, indent=2))


if __name__ == "__main__":
    sys.exit(main())
